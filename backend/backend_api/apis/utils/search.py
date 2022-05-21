from .inform_get import *
import threading
import dns
from .score import *
from dns import resolver
import findcdn
import re
import multiprocessing
import random
from collections import defaultdict
from bs4 import BeautifulSoup
import requests
import pymongo
import json
import os
import whois
from concurrent.futures import ThreadPoolExecutor, as_completed, ProcessPoolExecutor,wait,FIRST_COMPLETED,ALL_COMPLETED
import queue
import time
from threading import RLock
import sys
import openpyxl
from numpy import mean
base_path = "分类域名.xlsx"  #项目存放读取文件的路径
sys.path.append(base_path)
corpor_domains_average_dict = {'all_average': 0.5613793103448276, 'tld_average': 0.20689655172413793,
                               'register_average': -8.0, 'IP_average': 9.0, 'NS_average': 7.143720238095236,
                               'CDN_average': 0.5, 'cname_average': -1.262543859649123}
edu_domains_average_dict = {'all_average': 2.497586206896552, 'tld_average': 9.0, 'register_average': -8.0,
                            'IP_average': 9.0, 'NS_average': 8.44, 'CDN_average': 0,
                            'cname_average': 0.5444444444444443}
govern_domains_average_dict = {'all_average': 2.360344827586207, 'tld_average': 7.827586206896552,
                               'register_average': -8.0, 'IP_average': 9.0, 'NS_average': 7.606839080459769,
                               'CDN_average': 9.0, 'cname_average': 0.2717313762626261}
entertain_domains_average_dict = {'all_average': -0.12300000000000004, 'tld_average': -7.433333333333334,
                                  'register_average': -8.0, 'IP_average': 9.0, 'NS_average': 8.409472222222222,
                                  'CDN_average': 9.0, 'cname_average': 0.37993589743589745}
with open('NS_dict.json', 'r+') as file:
    NS_dict = file.read()
NS_dict = json.loads(NS_dict)  # 将json格式文件转化为python的字典文件
# 字典保存所有国内NS服务商所对应的国籍,不用看
with open('register_dict.json', 'r+') as file:
    register_dict = file.read()
register_dict = json.loads(register_dict)
with open('CDNdict.json', 'r+') as file:
    CDNdict = file.read()
CDNdict = json.loads(CDNdict)

# 字典保存了注册商与国家的对应关系
class cdn_Thread(threading.Thread):
    def __init__(self,domainname,q):
        threading.Thread.__init__(self)
        self.q = q
        self.domainname = domainname
        self.cdndict = {}
    def run(self):
        q = time
        self.cdndict = get_cdns(self.domainname)
        self.q.put(self.cdndict)

class HandleExcel:
    def load_excel(self):
        '''
        加载excel
        '''
        open_excel = openpyxl.load_workbook(base_path)#拿到excel的所有内容
        return open_excel
    def get_sheet_data(self,index=None):
        '''
        加载所有sheet的内容
        '''
        sheet_name = self.load_excel().sheetnames#拿到sheetnames的所有内容
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data
    def get_cell_value(self,row,cols):
        '''
        获取某一个单元格内容
        '''
        data = self.get_sheet_data().cell(row=row,column=cols)
        return data
    def get_rows(self):
        row = self.get_sheet_data().max_row
        return row
    def get_rows_value(self,row):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list


def operating_data_mongodb():
    try:
        con = pymongo.MongoClient('mongodb://root:#HITnist327@10.245.146.43:27077')
        return con
    except Exception as e:
        print(e)


def judge_legal_ip(ip):
    """判断一个IP地址是否合法"""
    compile_ip = re.compile('^((25[0-5]|2[0-4]\d|[01]?\d\d?)\.){3}(25[0-5]|2[0-4]\d|[01]?\d\d?)$')
    if compile_ip.match(ip):
        return True
    else:
        return False


# 下面三个函数用于过滤输入域名，在之后的函数编写中会用到
def get_standerd_domain_name(domainname):
    try:
        pattern = re.compile(r'[a-zA-Z0-9\-]+\.[a-zA-Z0-9\-]+\.[a-zA-Z0-9\-]+$')
        geted = (pattern.findall(domainname)[0]).lower()
    except:
        pattern = re.compile(r'[a-zA-Z0-9\-]+\.[a-zA-Z0-9\-]+$')
        geted = "www." + (pattern.findall(domainname)[0]).lower()
    return geted


def get_maindomain_name(domainname):
    pattern = re.compile(r'[a-zA-Z0-9\-]+\.[a-zA-Z0-9\-]+$')
    geted = (pattern.findall(domainname)[0]).lower()
    return geted


def get_mid_of_domain_name(domainname):
    pattern = re.compile(r'\.[a-zA-Z0-9\-]+\.')
    geted = (pattern.findall(domainname)[0]).lower()
    geted = geted[1:-1]
    return geted


def get_register_country(domain):
    geted = whois.whois("baidu.com")
    register = geted.get("registrar")
    try:
        register_country = register_dict.get(register)
    except:
        register_country = "NULL"
    return register_country


def get_tld_of_domain(domain):
    pattern = re.compile(r'\.[a-zA-Z0-9\-]+$')
    geted = (pattern.findall(domain)[0]).lower()
    return geted


def match_nscountry_by_org(domainname):  # 输入一个NS服务器域名，在字典NS_dict中匹配其主域名对应的所属国家打印ns及其国别
    country = NS_dict.get(get_maindomain_name(get_standerd_domain_name(domainname)))
    return country


def get_dns_and_layer(
        domain_name):  # 输入一个域名，获取其全部NS服务商的域名以及域名间的层级关系,存入数据库，需要设置存入的表（save_datacoll),每条解析出的NS# 域名做为一条记录.存入的格式为“domainname”字段对应输入的域名，“NSname”对应改NS服务器的域名，“NSlevel”对应改ns服务器的层级，#“preofNS”对应该ns服务器负责解析的域名（若level为1则对应解析输入的域名，其余解析的域名为上一层的NS服务器域名
    ns_inform_list = []
    standerd_domain_name = get_standerd_domain_name(domain_name)
    main_domain_name = get_maindomain_name(standerd_domain_name)
    try:
        ns1 = dns.resolver.resolve(main_domain_name,'NS')
        for i in ns1.response.answer:
            for j in i.items:
                j = str(j)
                j = j[:-1]
                ns1_main_domain_name = get_maindomain_name(j)
                list1 = []
                list1.append(domain_name)
                list1.append(j)
                list1.append(1)
                ns_inform_list.append(list1)
                # print(domain_name + "的第一级NS服务器有：" + j)
                if (get_mid_of_domain_name(j) != get_mid_of_domain_name(domain_name)):
                    ns2 = dns.resolver.resolve(ns1_main_domain_name, 'NS')
                    for p in ns2.response.answer:
                        for q in p.items:
                            q = str(q)
                            q = q[:-1]
                            if ((get_mid_of_domain_name(q) != get_mid_of_domain_name(j)) and (
                                    get_mid_of_domain_name(q) != get_mid_of_domain_name(domain_name))):
                                ns2_main_domain_name = get_maindomain_name(q)
                                list2 = []
                                list2.append(j)
                                list2.append(q)
                                list2.append(2)
                                ns_inform_list.append(list2)
                                # print(domain_name + "的第二级NS服务器有：" + q)
                                ns3 = dns.resolver.resolve(ns2_main_domain_name, 'NS')
                                for m in ns3.response.answer:
                                    for n in m.items:
                                        n = str(n)
                                        n = n[:-1]
                                        if (get_mid_of_domain_name(n) != get_mid_of_domain_name(q) and (
                                                get_mid_of_domain_name(n) != get_mid_of_domain_name(
                                                domain_name)) and get_mid_of_domain_name(n) != get_mid_of_domain_name(
                                                j)):
                                            ns3_main_domainname = get_maindomain_name(n)
                                            list3 = []
                                            list3.append(q)
                                            list3.append(n)
                                            list3.append(3)
                                            ns_inform_list.append(list3)
                                            # print(domain_name + "的第三级NS服务器有：" + n)
                                            ns4 = dns.resolver.resolve(ns3_main_domainname, 'NS')
                                            for a in ns4.response.answer:
                                                for b in a.items:
                                                    b = str(b)
                                                    b = b[:-1]
                                                    if (get_mid_of_domain_name(b) != get_mid_of_domain_name(
                                                            n) and get_mid_of_domain_name(b) != get_mid_of_domain_name(
                                                            q) and (get_mid_of_domain_name(b) != get_mid_of_domain_name(
                                                            domain_name)) and get_mid_of_domain_name(
                                                            b) != get_mid_of_domain_name(j)):
                                                        ns4_main_domainname = get_maindomain_name(b)
                                                        list4 = []
                                                        list4.append(n)
                                                        list4.append(b)
                                                        list4.append(4)
                                                        ns_inform_list.append(list4)
                                                        # print(domain_name+"的第四级NS服务器有："+b)
                                                        ns5 = dns.resolver.resolve(ns4_main_domainname)
                                                        for x in ns5.response.answer:
                                                            for y in x.items:
                                                                y = str(y)
                                                                y = y[:-1]
                                                                if (get_mid_of_domain_name(y) != get_mid_of_domain_name(
                                                                        b)):
                                                                    list5 = []
                                                                    list5.append(b)
                                                                    list5.append(y)
                                                                    list5.append(5)
                                                                    ns_inform_list.append(list5)
                                                                # print(domain_name + "的第五级NS服务器有:"+ y)
    except Exception as e:
        print("failed", e)
    return ns_inform_list


# 下面是获取域名的物理地址

def GetArecordIp(domain_name):  # 输入一个域名，获取它的IPV4地址，返回的是一个列表
    address = []
    try:
        host_a = dns.resolver.resolve(domain_name, 'A')
        for i in host_a.response.answer:
            for j in i.items:
                j = str(j)
                if (judge_legal_ip(j)):
                    address.append(j)
        return address  # 这是返回的列表
    except:
        return []


def get_country(ip_addr):  # 输入一个IP地址返回一个国家
    try:
        ip = ip_addr
        res = requests.get("http://pypi.hitwh.net.cn:7788/ip/info?ip=" + ip + "&acc=city")
        ip_res = json.loads(res.text)["data"]
        return ip_res.get("country")
    except:
        return "NULL"


def get_country_by_ip(domain):  # 输入一个域名，获取国家与地址，并且存入相应的数据库，即组合了上面两个函数
    addrlist = GetArecordIp(domain)
    countrylist = []
    for address in addrlist:
        country = get_country(address)
        countrylist.append(country)
    return countrylist




def get_tld_domain_details(domain):  # 获取并打印域名顶级域相关信息,返回一个信息存储字典
    """
    获取域名详情
    :param domain:
    :return: dictuionary: 详情
    """
    domain = get_tld_of_domain(domain)
    url = "https://www.iana.org/domains/root/db/" + domain[1:] + ".html"
    try:
        res = requests.get(url)
        soup = BeautifulSoup(res.text, 'lxml')
        message = soup.select('.hemmed')[0]  # 所有数据皆集中于此
    except Exception as e:
        return domain + "根域名过于奇怪"
    h2s = message.select('h2')
    pattern = r"<h2>"
    sibs = []
    try:
        for sibling in h2s[0].next_siblings:
            if len(re.findall(pattern, repr(sibling))):
                break
            if sibling.string == None or sibling.string == '\n':
                continue
            if len(re.findall(r"\\n", repr(sibling))):
                sibling.string = sibling.string[5:]
            sibs.append(sibling.string)
        add = " ".join(sibs[1:-1])
        orz = sibs[0]  # 主办单位
        country = sibs[len(sibs) - 1]  # 国家
        zip_code = sibs[len(sibs) - 2][sibs[len(sibs) - 2].rfind(" ") + 1:]  # 邮编
        address = add[:add.rfind(" ")]  # 地址
    except Exception as e:
        return {
            '顶级域名': domain,
            '主办单位': "不详",
            '国家': "不详",
            '地址': "不详",
            '邮编': "不详"
        }
    info = {
        '顶级域名': domain,
        '主办单位': orz,
        '国家': country,
        '地址': address,
        '邮编': zip_code
    }
    return {
        '顶级域名': domain,
        '主办单位': orz,
        '国家': country,
        '地址': address,
        '邮编': zip_code
    }


def get_cdns(domain):  # 输入一个域名列表，即可获取列表中域名的CDN信息，并打印相关信息
    domain_list = []
    domain_list.append(domain)
    i = get_standerd_domain_name(domain)
    path = 'output'+i+'.json'
    resp_json = findcdn.main(domain_list=domain_list, output_path='output'+i+'.json', double_in=True, threads=23,timeout=10)
    dumped_json = json.loads(resp_json)  # 运行后会在每次获取之后生成一个json文件（没啥用）
      # 删除生成的json文件
    if os.path.exists(path):
        os.remove(path)
    domainname = domain  # 需要获取CDN信息的域名
    cdnname = []  # 得到与域名相关的CDN服务商的列表
    try:
        cdnnameget = eval(dumped_json['domains'][domain]['cdns_by_names'])
        cdnname.append(cdnnameget)
        cdncountry = []
        for i in cdnname:  # 对列表中的每个CDN服务商，在字典中寻找其国籍，并依次存入cdncountry列表
            cdncountry.append(CDNdict.get(i))
        return {"cdn服务商": cdnname, "对应的国家": cdncountry}
    except:
        return {"cdn服务商": ["NULL"], "对应的国家": ["NULL"]}


def getCnamerember(domain):  # 获取域名cname记录
    """
    仅限输入二级域名，否则出错
    :param domain:
    :return: 打印该域名的cname记录值
    """
    cname_list = []
    try:
        cname = dns.resolver.resolve(domain, 'CNAME')
        for i in cname.response.answer:
            for j in i.items:
                j = j.to_text()
                j = j[:-1]
                cname_list.append(j)
    except Exception as e:
        print(e)
    return cname_list


def getArember(domain):
    """
    仅查询二级域名，顶级域名则出错
    :param domain:
    :return:打印该域名下的所有A记录
    """
    A = dns.resolver.query(domain, 'A')  # 指定查询A记录
    ip_list = []
    for i in A.response.answer:  # 通过response.answer方法获取查询回应的信息
        for j in i:
            ip_list.append(j)
    return ip_list


def get_ns_inform_list(domain):
    ns_inform_list = get_dns_and_layer(domain)
    ns_dict_list = []  # 列表存有关于该域名的NS信息
    for i in ns_inform_list:  # 每一条相关的NS解析记录记为一个字典,最终暂存于列表ns_dict_list中
        resloved_domain = i[0]
        ns_domain_name = i[1]
        level = i[2]
        nscountry_by_ip = get_country_by_ip(ns_domain_name)
        nscountry_by_org = match_nscountry_by_org(ns_domain_name)
        ns_dict = {"域名": resloved_domain, "解析它的NS服务器域名": ns_domain_name, "NS服务器所处国家": nscountry_by_ip,
                   "NS服务器所属国家": nscountry_by_org, "解析层级": level}
        ns_dict_list.append(ns_dict)
    return ns_dict_list


def get_cname_register_inform(cname_list):
    cname_registers = []  # 列表，存放了字典，即cname与其注册商国别的对应关系
    for cname in cname_list:
        cname_registers.append({cname: get_register_country(cname)})
    return cname_registers


def get_cname_ns_inform(cname_list):
    cname_ns_inform = []  # 二维列表，其中的每个列表存放着一个cname的所有ns信息,信息是以字典的形式保存的
    for cname in cname_list:
        one_cname_dict_list = []
        cname_inform_list = get_dns_and_layer(cname)
        for j in cname_inform_list:
            resloved_domain = j[0]
            ns_domain_name = j[1]
            level = j[2]
            nscountry_by_ip = get_country_by_ip(ns_domain_name)
            nscountry_by_org = match_nscountry_by_org(ns_domain_name)
            ns_of_cname_dict = {"被解析的域名": resloved_domain, "解析它的NS服务器域名": ns_domain_name, "NS服务器所处国家": nscountry_by_ip,
                                "NS服务器所属国家": nscountry_by_org, "解析层级": level}
            one_cname_dict_list.append(ns_of_cname_dict)
        cname_ns_inform.append({cname: one_cname_dict_list})
    return cname_ns_inform

def find_inform_of_one_domain(domain , collname):
    con = operating_data_mongodb()
    db = con["domain_country"]
    coll_name = db[collname]
    finded = coll_name.find({"域名":domain})
    if list(coll_name.find({"域名":domain})) == []:
        return 0
    else:
        return finded[0]

def alexa_search_of_one_domain(domain):
    key ="69915ef10de44c419fb594eacb6f5b8a"
    res = requests.get("https://apidatav2.chinaz.com/single/alexa?key="+key+"&domain="+domain)
    ip_res = json.loads(res.text)
    result = ip_res.get("Result")
    IpNum = result.get("IpNum")
    keyword = result.get("Keywords")
    return IpNum

def get_message(data):
    flag = 1
    for i in range(1,6):
        if data[i] < 0:
            flag = 0
            break
    if flag == 1:
        message = "不存在风险"
    else:
        message = "存在一定风险"
    return message

def get_detail_analyse(data,visit_count,type,score_dict,message):
    if (type == "1"):  # 教育类型
        avescore = edu_domains_average_dict.get("all_average")
        if visit_count > 5000000:
            comprehansive_impact = "很高"
        elif visit_count > 1000000:
            comprehansive_impact = "较高"
        elif visit_count > 100000:
            comprehansive_impact = "一般"
        else:
            comprehansive_impact = "较低"

    elif (type == "2"):  # 政府机关类型
        avescore = govern_domains_average_dict.get("all_average")
        if visit_count > 5000000:
            comprehansive_impact = "很高"
        elif visit_count > 1000000:
            comprehansive_impact = "较高"
        elif visit_count > 100000:
            comprehansive_impact = "一般"
        else:
            comprehansive_impact = "较低"
    elif (type == "3"):  # 银行/公司官网类型
        avescore = corpor_domains_average_dict.get("all_average")
        if visit_count > 5000000:
            comprehansive_impact = "很高"
        elif visit_count > 1000000:
            comprehansive_impact = "较高"
        elif visit_count > 100000:
            comprehansive_impact = "一般"
        else:
            comprehansive_impact = "较低"
    else:  # 娱乐购物类网站
        avescore = entertain_domains_average_dict.get("all_average")
        if visit_count > 5000000:
            comprehansive_impact = "很高"
        elif visit_count > 1000000:
            comprehansive_impact = "较高"
        elif visit_count > 100000:
            comprehansive_impact = "一般"
        else:
            comprehansive_impact = "较低"
    avescore = round(avescore,2)
    if message == "不存在风险":
        final_list = [comprehansive_impact, avescore]
        # 列表中顺序与此文字中通配符顺序对应“该域名综合影响力%s,该类型域名平均得分为%d,从域名解析背后实体的国家属性来看此域名是安全的"
        return final_list
    else:
        data = data[1:]
        mi = min(data)
        j = 0
        for i in data:
            if i == mi:
                break
            else:
                j = j+1
        #注册商，IP，NS，CDN，别名
        if j == 0:
            min_name = "注册商"
        elif j == 1:
            min_name = "IP物理地址"
        elif j == 2:
            min_name = "NS服务商"
        elif j == 3:
            min_name = "CDN服务商"
        elif j == 4:
            min_name = "别名"
        if data[j] == 0:
            second_min = max(data)
            for j in range(0,len(data)):
                if data[j] != min:
                    if data[j] < second_min:
                        second_min = data[j]
            j = 0
            for i in data:
                if i == second_min:
                    break
                else:
                    j = j + 1
            if j == 0:
                min_name = "注册商"

            elif j == 1:
                min_name = "IP物理地址"

            elif j == 2:
                min_name = "NS服务商"

            elif j == 3:
                min_name = "CDN服务商"

            elif j == 4:
                min_name = "别名"

        if min_name == "注册商":
            score_dict.update({"注册商得分": 9})
            advise = "将域名迁移至中国注册商，如新网、万网、商务中国的管理下"
        elif min_name == "IP对应物理地址":
            score_dict.update({"IP对应物理地址得分": 9})
            advise = "采用国内的服务器来运行此域名对应网站，较好的服务器提供商有阿里云，腾讯云，华为云等"
        elif min_name == "NS服务商":
            score_dict.update({"NS信息得分": 9})
            advise = "建议选用国内的NS服务商来对此域名展开解析，如DNSPOD.CN，Aidns.cn，Everdns.cn"
        elif min_name == "CDN服务商":
            score_dict.update({"CDN得分": 9})
            advise = "建议选用国内的CDN服务商,如阿里云CDN，腾讯云CDN等"
        else :
            score_dict.update({"别名得分": 9})
            advise = "建议更改此域名别名的各项属性，并尽可能使用属于中国的服务商与注册商"
        changed_score = score_for_one_domain(score_dict)
        final_list = [comprehansive_impact,avescore,min_name,advise,changed_score]
        #列表中顺序与此文字中通配符顺序对应“该域名综合影响力%s,该类型域名平均得分为%d,此域名%s一项得分最低，建议%s,修改后域名的总体得分为%s“
        return final_list



def run_one(domain,type):
    m = find_inform_of_one_domain(domain,"try")
    if m == 0:
        q = queue.Queue()
        cdn_dict = {}
        subthread = cdn_Thread(domain, q)
        subthread.start()
        tld_country = get_tld_domain_details(domain).get("国家")  # 获得顶级域国籍
        register_country = get_register_country(domain)  # 获得注册商国籍
        ip_list = GetArecordIp(domain)  # 获得IP地址
        ip_country_list = []  # 获得IP地址所处国家
        for ip in ip_list:
            ip_country = get_country(ip)
            ip_country_list.append(ip_country)
        ns_dict_list = get_ns_inform_list(domain)  # 获取全部ns信息，为一个列表，里面单元为字典
        cname_list = getCnamerember(domain)  # 获取别名信息,列表
        if cname_list is not None:
            cname_registers = get_cname_register_inform(cname_list)  # 列表，存放了字典，即cname与其注册商国别的对应关系
            cname_ns_inform = get_cname_ns_inform(cname_list)  # 二维列表，其中的每个列表存放着一个cname的所有ns信息,信息是以字典的形式保存的
        else:
            cname_list = ["NULL"]
            cname_registers = ["NULL"]
            cname_ns_inform = ["NULL"]
        subthread.join()
        cdn_dict = q.get()
        cdn_server = cdn_dict.get("cdn服务商")  # 列表,存储cdn服务商
        cdn_country = cdn_dict.get("对应的国家")  # 列表，存储cdn服务商对应的国家
        final_inform = {"域名": domain, "顶级域所属国家": tld_country, "注册商所属国家": register_country, "IP地址": ip_list,
                        "IP所处国家": ip_country_list,
                        "NS服务器信息": ns_dict_list, "别名": cname_list, "别名对应注册商国家": cname_registers,
                        "别名的NS服务器信息": cname_ns_inform,"CDN服务商信息": cdn_dict
                        }
    else:
        del m['_id']
        final_inform = m
        ns_dict_list = m.get("NS服务器信息")
        ip_list = m.get("IP地址")
        cname_list = m.get("别名")
        cdn_server = m.get("CDN服务商信息").get("cdn服务商")
    score_dict = inform_to_score(domain,final_inform)
    score_dict = score_dict.get(domain)
    final_score = score_for_one_domain(score_dict)
    print(domain + "inform geted successfully")
    nsserverlist = []
    for item in  ns_dict_list:
        if item.get("域名") == domain:
            name = get_maindomain_name(get_standerd_domain_name(item.get("解析它的NS服务器域名")))
            if name not in nsserverlist:
                nsserverlist.append(name)
    data = []
    avedata = []
    for iten in score_dict:
        data.append(score_dict.get(iten))#列表顺序为顶级域，注册商，IP，NS，CDN，别名
    data = data[1:]
    if(type == "1"):   #教育类型
        for iten in edu_domains_average_dict:
            avedata.append(edu_domains_average_dict.get(iten))
        avedata = avedata[1:]#列表顺序为顶级域，注册商，IP，NS，CDN，别名
    elif(type== "2"):  #政府机关类型
        for iten in govern_domains_average_dict:
            avedata.append(govern_domains_average_dict.get(iten))
        avedata = avedata[1:]#列表顺序为顶级域，注册商，IP，NS，CDN，别名
    elif(type == "3"):  #银行/公司官网类型
        for iten in corpor_domains_average_dict:
            avedata.append(corpor_domains_average_dict.get(iten))
        avedata = avedata[1:]#列表顺序为顶级域，注册商，IP，NS，CDN，别名
    else:  #娱乐购物类网站
        for iten in entertain_domains_average_dict:
            avedata.append(entertain_domains_average_dict.get(iten))
        avedata = avedata[1:]  # 列表顺序为顶级域，注册商，IP，NS，CDN，别名
    for i in range(len(data)):
        data[i] = round(data[i],2)
    for j in range(len(avedata)):
        avedata[j] = round(avedata[j],2)
    visit_count = alexa_search_of_one_domain(domain)
    message = get_message(data)
    detailanalyse = get_detail_analyse(data,visit_count,type,score_dict,message)
    IP = "、".join(ip_list)
    cname = "、".join(cname_list)
    nsserver = "、".join(nsserverlist)
    cdnserver = "、".join(cdn_server)
    final_dict = {"domainName":domain,"IP":IP,"cname":cname,
                  "NSServer":nsserver,"CDNServer":cdnserver,"visit_count":visit_count ,
                  "point":final_score,"pointmessage":message,"data":data,"avedata":avedata,"detailanalyse":detailanalyse}
    return final_dict

def run_one_listuse(canshu_list):
    domain = canshu_list[0]
    m = find_inform_of_one_domain(domain,"try")
    if m == 0:
        tld_country = get_tld_domain_details(domain).get("国家")  # 获得顶级域国籍
        register_country = get_register_country(domain)  # 获得注册商国籍
        ip_list = GetArecordIp(domain)  # 获得IP地址
        ip_country_list = []  # 获得IP地址所处国家
        for ip in ip_list:
            ip_country = get_country(ip)
            ip_country_list.append(ip_country)
        ns_dict_list = get_ns_inform_list(domain)  # 获取全部ns信息，为一个列表，里面单元为字典
        cname_list = getCnamerember(domain)  # 获取别名信息,列表
        if cname_list is not None:
            cname_registers = get_cname_register_inform(cname_list)  # 列表，存放了字典，即cname与其注册商国别的对应关系
            cname_ns_inform = get_cname_ns_inform(cname_list)  # 二维列表，其中的每个列表存放着一个cname的所有ns信息,信息是以字典的形式保存的
        else:
            cname_list = ["NULL"]
            cname_registers = ["NULL"]
            cname_ns_inform = ["NULL"]
        cdn_dict = get_cdns(domain)
        final_inform = {"域名": domain, "顶级域所属国家": tld_country, "注册商所属国家": register_country, "IP地址": ip_list,
                        "IP所处国家": ip_country_list,
                        "NS服务器信息": ns_dict_list, "别名": cname_list, "别名对应注册商国家": cname_registers,
                        "别名的NS服务器信息": cname_ns_inform,"CDN服务商信息": cdn_dict
                        }
    else:
        del m['_id']
        final_inform = m
        ns_dict_list = m.get("NS服务器信息")
        ip_list = m.get("IP地址")
        cname_list = m.get("别名")
        cdn_server = m.get("CDN服务商信息").get("cdn服务商")
    print(domain + "inform geted successfully")
    if canshu_list[2] == 1:
        newcon = pymongo.MongoClient('mongodb://root:#HITnist327@10.245.146.43:27077')
        db = newcon['domain_country']
        save_col = db[canshu_list[1]]
        save_col.insert_one(final_inform)
        newcon.close()

        print(domain + "inform geted successfully")
    else:
        pass

def e_run(canshulist):
    sublist = canshulist[0]
    coll = canshulist[1]
    if_insert = canshulist[2]
    with ThreadPoolExecutor(max_workers=10) as threadpool:
        all_task = []
        for domain in sublist:
            canshu = [domain, coll, if_insert]
            all_task.append(threadpool.submit(run_one_listuse, canshu))
            wait(all_task,timeout= 200, return_when=ALL_COMPLETED)
    threadpool.shutdown(True)


def runlist_to_db(domainlist,db_name,collname,process,if_insert):
    dbcon = operating_data_mongodb()
    database = dbcon[db_name]
    coll = database[collname]
    whole_list = []
    start = 0
    sum = len(domainlist)
    if (sum % process == 0):
        per_len = sum//process
    else:
        per_len = sum//process + 1
    for i in range(0,process):
        sublist = domainlist[start:start+per_len]
        whole_list.append(sublist)
        start = start+per_len
    joincanshu = []
    for item in whole_list:
        l = [item,collname,if_insert]
        joincanshu.append(l)
    pool = multiprocessing.Pool(processes=process)
    pool.map(e_run,joincanshu)
    pool.close()
    pool.join()

def get_domainname_list(datacoll, start, end):
    finded = datacoll.find({})
    whole_list = []
    for i in finded:
        whole_list.append(i.get('domain'))
    sublist1 = whole_list[start:start + int((end - start) / 2)]
    sublist2 = whole_list[start + int((end - start) / 2):end]
    list = [sublist1, sublist2]
    return list

def get_type_of_domian_inform(collname):
    con = operating_data_mongodb()
    db = con["domain_country"]
    data_coll = db[collname]
    finded = data_coll.find({})
    tld_score_list = []
    register_score_list = []
    IP_score_list = []
    NS_score_list = []
    CDN_score_list = []
    cname_score_list = []
    final_score_list = []
    for i in finded:
        first_score_dict = inform_to_score(i.get("域名"),i)
        first_score_dict = first_score_dict.get(i.get("域名"))
        total_score = score_for_one_domain(first_score_dict)
        final_score_list.append(total_score)
        tld_score = first_score_dict.get('顶级域得分')
        register_score = first_score_dict.get('注册商得分')
        IP_score = first_score_dict.get('IP对应物理地址得分')
        NS_score = first_score_dict.get('NS信息得分')
        CDN_score = first_score_dict.get('CDN得分')
        cname_score = first_score_dict.get('别名得分')
        if (tld_score != 0):
            tld_score_list.append(tld_score)
        if (register_score != 0):
            register_score_list.append(register_score)
        if (IP_score != 0):
            IP_score_list.append(IP_score)
        if (NS_score != 0):
            if (NS_score >= 9):
                NS_score_list.append(9)
            elif (NS_score <= -8):
                NS_score_list.append(-8)
            else:
                NS_score_list.append(NS_score)
        if (CDN_score != 0):
            CDN_score_list.append(CDN_score)
        if (cname_score != 0):
            if (cname_score >= 9):
                cname_score_list.append(9)
            elif (cname_score <= -8):
                cname_score_list.append(-8)
            else:
                cname_score_list.append(cname_score)

    final_dict = {"all_average":mean(final_score_list),"tld_average":mean(tld_score_list),"register_average":mean(register_score_list),
                 "IP_average":mean(IP_score_list),"NS_average":mean(NS_score_list),"CDN_average":mean(CDN_score_list),"cname_average":mean(cname_score_list)
                 }
    return final_dict

def get_time(fmt:str='%Y-%m-%d %H-%M-%S') -> str:
    '''
    获取当前时间
    '''
    ts = time.time()
    ta = time.localtime(ts)
    t = time.strftime(fmt, ta)
    return t

def multifind(domain_list):
    final_list = []
    dbname = "domain_country"
    collname = get_time()
    runlist_to_db(domain_list, dbname, collname, 6, 1)
    new_con = operating_data_mongodb()
    db = new_con[dbname]
    datacoll = db[collname]
    finded = datacoll.find({})
    for i in finded:
        final_inform = i
        del final_inform['_id']
        domain = final_inform.get("域名")
        ip_list = final_inform.get("IP地址")
        score_dict = inform_to_score(domain,final_inform)
        score_dict = score_dict.get(domain)
        final_score = score_for_one_domain(score_dict)
        IP = "、".join(ip_list)
        data = []
        for iten in score_dict:
            data.append(score_dict.get(iten))  # 列表顺序为顶级域，注册商，IP，NS，CDN，别名
        data = data[1:]
        for bc in range(len(data)):
            data[bc] = round(data[bc],2)
        subdict = {"domainName":domain,"IP":IP,"point":final_score,"data":data}
        final_list.append(subdict)
    return final_list
if __name__ == '__main__':
    domainlist = ["www.baidu.com","www.sohu.com","www.qq.com","www.hteacher.net","www.leiphone.com","www.126.com"]
    print(multifind(domainlist))

    '''
    使用multifind（）函数完成对于列表的查询，返回值为一个列表，列表的各项是各域名对应信息的字典
    '''
    # print(run_one("www.baidu.com",1))

    '''
    type = 1 对应教育类 ，2 对应政府机关类，3对应银行/企业官网类，4对应娱乐/购物网站类
    run_one（domain,type) 用于单域名查询其返回值为一个字典，样例如下
    {'domainName': 'www.sohu.com', 'IP': '221.179.177.18', 'cname': 'gs.a.sohu.com', 'NSServer': 'sohu.com',
     'CDNSercer': 'NULL', 'visit_count': 4290000, 'point': -0.57,
      'pointmessage': '不存在风险', 'data': [-8, -8, 9.0, 9.0, 0.0, 0.5], 'avedata': [-7.43, -8.0, 9.0, 8.41, 9.0, 0.38],
       'detailanalyse': ['较高', -0.12]}
    data 与 avedata 的顺序都是[顶级域得分，注册商得分，IP对应物理地址得分，NS得分，CDN得分，别名得分]
    当'pointmessage' == 不存在风险时，
    detailanalyse列表有2项，对应的分析为“该域名综合影响力%s,该类型域名平均得分为%d,从域名解析背后实体的国家属性来看此域名是安全的”
    当'pointmessage' == 存在风险时，
    detailanalyse列表有5项，对应的分析为“该域名综合影响力%s,该类型域名平均得分为%d,此域名%s一项得分最低，建议%s,修改后域名的总体得分为%s”
    '''























